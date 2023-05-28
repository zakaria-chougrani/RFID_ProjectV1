from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models import Employe


def export_csv(request, pk):
    employe = get_object_or_404(Employe, pk=pk)
    events = employe.entree_sorties.filter(employe=employe).order_by("-id")

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Nom:'
    ws['B1'] = employe.nom
    ws['A2'] = 'Prénom:'
    ws['B2'] = employe.prenom
    ws['A3'] = 'RFID N°:'
    ws['B3'] = employe.num_rfid
    ws['A4'] = 'ID EVENT'
    ws['B4'] = 'TYPE'
    ws['C4'] = 'DATE'
    ws['D4'] = 'TIME'
    # Ajoutez les données à la feuille
    for i, row in enumerate(events):
        ws.cell(row=i + 5, column=1, value=row.id)
        ws.cell(row=i + 5, column=2, value=row.type_event)
        ws.cell(row=i + 5, column=3, value=row.date_event)
        ws.cell(row=i + 5, column=4, value=row.time)

    # set column widths
    column_widths = [10, 20, 15, 15]  # set desired column widths
    for i, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + i)].width = width  # adjust column width based on index
    # format the data as a table
    tab = Table(displayName="employee_events", ref="A4:D" + str(len(events) + 4))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Renvoyez le fichier Excel en réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_events.xlsx'
    wb.save(response)
    return response
